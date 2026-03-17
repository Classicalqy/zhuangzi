#!/usr/bin/env python3
from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Dict, List, Literal, Tuple

import openpyxl


Kind = Literal["annotation", "interpretation"]


@dataclass(frozen=True)
class TextConfig:
    sheet: str
    header_row: int
    data_start: int
    text_id: int
    text_title: str


TEXT_CONFIGS: List[TextConfig] = [
    TextConfig("Sheet1", 1, 2, 1, "人间世"),
    TextConfig("Sheet2", 2, 3, 2, "逍遥游"),
    TextConfig("Sheet3", 1, 2, 3, "养生主"),
]

ANNOTATION_BOOK_KEYS = ["音义"]
INTERPRETATION_BOOK_KEYS = ["注疏", "口义", "义海", "内篇注", "庄子解", "经解", "集释", "集解", "庄子翼"]

ANNOTATION_COMMENTATORS = {
    "陆德明",
    "司马彪",
    "郭象",
    "崔撰",
    "向秀",
    "李颐",
    "李轨",
    "徐邈",
    "萧纲（梁简文帝）",
    "梁简文帝",
    "简文",
    "嵇康",
    "东方朔《十洲记》",
    "郭璞",
    "支遁",
    "潘尼",
    "《楚辞天问》王逸注",
    "王逸注",
    "方以智",
    "卢文弨",
    "郭庆藩",
    "王先谦",
    "姚鼐",
    "郭嵩焘",
    "苏舆",
    "俞樾",
    "李桢",
    "王念孙",
}

INTERPRETATION_COMMENTATORS = {
    "郭象（注）",
    "成玄英（疏）",
    "成玄英",
    "林希逸",
    "吕惠卿",
    "林疑獨",
    "陳詳道",
    "陈详道",
    "陳景元（碧虚）",
    "趙以夫",
    "褚伯秀",
    "释德清",
    "王夫之",
    "宣颖",
    "王旦",
    "王雱",
}

# Strict annotation markers (avoid broad single-char matches that cause false positives).
PRON_RE = re.compile(r"(?:^|[，。；\s])音[\u4e00-\u9fffA-Za-z]{1,6}")
FANQIE_RE = re.compile(r"[\u4e00-\u9fff]{1,3}[，,][^。；，,\n]{0,3}反(?:[。；，,\s]|$)")
VARIANT_RE = re.compile(r"本亦作|本作|又作|作“|作\"|读为|讀爲|读曰|讀曰|一作|或作|当作|當作|误作|誤作|假借为|假借作|借爲|借为")
DICT_RE = re.compile(r"字林|说文|説文|尔雅|爾雅|广雅|廣雅|玉篇|释文|釋文")
SAMENESS_RE = re.compile(r"下同|上同|并同|並同|一音")
NAME_KIND_RE = re.compile(r"书名|人名|地名|国名|國名|县名|縣名|官名")
PERSON_DEF_RE = re.compile(r"姓[^。；\n]{0,8}名[^。；\n]{0,8}(?:[^。；\n]{0,8}字[\u4e00-\u9fff]{1,4})?")
GLOSS_VERB_RE = re.compile(r"谓之|謂之|谓|謂|犹|猶|即")
HEAD_GLOSS_RE = re.compile(r"^[「“]?[^\s，,。；]{1,8}[」”]?(?:者)?[，,:：][^，,。；\n]{1,16}(?:也|者也|耳|矣)$")
INTERP_KEYWORDS = [
    "道",
    "德",
    "心",
    "性",
    "无为",
    "無爲",
    "至人",
    "圣人",
    "聖人",
    "天下",
    "阴阳",
    "陰陽",
    "逍遥",
    "逍遙",
    "自然",
    "善恶",
    "善惡",
    "名刑",
    "养生",
    "養生",
    "物",
    "理",
    "气",
    "氣",
    "神",
]
CORE_INTERP_KEYWORDS = ["道", "德", "心", "性", "无为", "無爲", "至人", "圣人", "聖人", "天下", "阴阳", "陰陽", "逍遥", "逍遙", "自然", "善恶", "善惡", "名刑", "养生", "養生"]

META_HEADER_LABELS = {"注者", "分工", "负责人", "朝代", "书名", "特点"}


def normalize(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").replace("\r", "\n")
    lines = [line.strip() for line in s.split("\n")]
    return "\n".join([line for line in lines if line]).strip()


def ann_score(text: str) -> int:
    score = 0
    if PRON_RE.search(text):
        score += 2
    if FANQIE_RE.search(text):
        score += 2
    if VARIANT_RE.search(text):
        score += 2
    if DICT_RE.search(text):
        score += 2
    if SAMENESS_RE.search(text):
        score += 1
    if NAME_KIND_RE.search(text):
        score += 2
    if PERSON_DEF_RE.search(text):
        score += 1
    if len(text) <= 14:
        score += 1
    if text.count("。") == 0 and text.count("，") <= 2 and len(text) <= 24:
        score += 1
    return score


def gloss_clause_count(text: str) -> int:
    clauses = [x.strip(" 「」“”\"'") for x in re.split(r"[。；;\n]", text) if x.strip()]
    count = 0
    for clause in clauses:
        if len(clause) <= 40 and HEAD_GLOSS_RE.match(clause):
            count += 1
            continue
        if len(clause) <= 28 and (PRON_RE.search(clause) or FANQIE_RE.search(clause)):
            count += 1
            continue
        if len(clause) <= 32 and (NAME_KIND_RE.search(clause) or PERSON_DEF_RE.search(clause)):
            count += 1
            continue
    return count


def has_hard_annotation_marker(text: str) -> bool:
    return bool(
        PRON_RE.search(text)
        or FANQIE_RE.search(text)
        or VARIANT_RE.search(text)
        or DICT_RE.search(text)
        or SAMENESS_RE.search(text)
        or NAME_KIND_RE.search(text)
        or PERSON_DEF_RE.search(text)
    )


def is_gloss_like_clause(text: str) -> bool:
    t = text.strip(" 「」“”\"'")
    if not t:
        return False
    if has_hard_annotation_marker(t):
        return True
    if len(t) <= 40 and HEAD_GLOSS_RE.match(t):
        return True
    if len(t) <= 22 and re.search(r"[^\s，,。；]{1,8}[，,:：][^，,。；]{1,14}(?:谓之|謂之|谓|謂|犹|猶|即)", t):
        return True
    return False


def split_cell_by_sentence(text: str) -> List[str]:
    """Coarse segmentation: keep sentence-level order, avoid over-splitting by commas."""
    if not text:
        return []

    parts = re.split(r"([。；;！？?!\n]+)", text)
    units: List[str] = []
    i = 0
    while i < len(parts):
        body = (parts[i] or "").strip()
        sep = ""
        if i + 1 < len(parts) and parts[i + 1]:
            sep = parts[i + 1]
        seg = f"{body}{sep}".strip()
        if seg:
            units.append(seg)
        i += 2

    return units if units else [text.strip()]


def interp_score(text: str) -> int:
    score = 0
    if len(text) >= 30:
        score += 2
    if len(text) >= 80:
        score += 2
    punct = text.count("。") + text.count("；") + text.count(";")
    if punct >= 2:
        score += 1
    kw_count = sum(1 for kw in INTERP_KEYWORDS if kw in text)
    if kw_count >= 1:
        score += 1
    if kw_count >= 3:
        score += 1
    if any(token in text for token in ("所以", "故", "盖", "蓋")):
        score += 1
    return score


def row_default_kind(sheet: str, row: int, book: str, commentator: str) -> Kind:
    if any(k in book for k in ANNOTATION_BOOK_KEYS):
        return "annotation"
    if any(k in book for k in INTERPRETATION_BOOK_KEYS):
        return "interpretation"
    if commentator in ANNOTATION_COMMENTATORS:
        return "annotation"
    if commentator in INTERPRETATION_COMMENTATORS:
        return "interpretation"

    # Fallback to observed split positions in the three source sheets.
    if sheet == "Sheet1":
        return "interpretation" if row >= 11 else "annotation"
    if sheet == "Sheet2":
        return "interpretation" if row >= 18 else "annotation"
    if sheet == "Sheet3":
        return "interpretation" if row >= 10 else "annotation"
    return "annotation"


def classify_cell(sheet: str, row: int, book: str, commentator: str, text: str) -> Kind:
    default = row_default_kind(sheet, row, book, commentator)
    a_score = ann_score(text)
    i_score = interp_score(text)
    hard_markers = {
        "pron": bool(PRON_RE.search(text)),
        "fanqie": bool(FANQIE_RE.search(text)),
        "variant": bool(VARIANT_RE.search(text)),
        "dict": bool(DICT_RE.search(text)),
        "same": bool(SAMENESS_RE.search(text)),
        "name_kind": bool(NAME_KIND_RE.search(text)),
        "person_def": bool(PERSON_DEF_RE.search(text)),
    }
    marker_hits = sum(1 for v in hard_markers.values() if v)
    has_hard_ann = marker_hits > 0
    core_kw_hits = sum(1 for kw in CORE_INTERP_KEYWORDS if kw in text)
    gloss_hits = gloss_clause_count(text)
    quote_book_hits = text.count("《")
    yun_hits = text.count("云")
    punct_hits = text.count("。") + text.count("；") + text.count(";") + text.count("，")
    text_len = len(text)

    # Annotation is strictly word-level/lexical explanation.
    if is_gloss_like_clause(text):
        return "annotation"
    if has_hard_ann and gloss_hits >= 1:
        return "annotation"
    if has_hard_ann and (hard_markers["pron"] or hard_markers["fanqie"] or hard_markers["variant"] or hard_markers["dict"]):
        return "annotation"
    if has_hard_ann and marker_hits >= 2 and core_kw_hits <= 1 and text_len < 180:
        return "annotation"
    # Loosen annotation slightly: short non-philosophical clauses tend to be lexical notes.
    if text_len <= 20 and core_kw_hits == 0 and i_score <= 1 and punct_hits <= 1:
        return "annotation"
    if re.search(r"^[^，,。；]{1,10}者[，,][^。；]{1,18}(?:也|者也|耳|矣)$", text):
        return "annotation"

    if commentator in ANNOTATION_COMMENTATORS and has_hard_ann:
        return "annotation"

    # 文本批评/音训密集，且缺少核心哲学语义 -> annotation
    if has_hard_ann and core_kw_hits == 0 and (marker_hits >= 2 or (quote_book_hits >= 2 and yun_hits >= 2)):
        return "annotation"

    # 句内多处短义训，基本属于 annotation。
    if gloss_hits >= 2 and core_kw_hits <= 2:
        return "annotation"

    # 极短文本大概率是注释项
    if len(text) <= 8 and i_score <= 1:
        return "annotation"

    # Interpretation now includes:
    # 1) philosophical reflection
    # 2) sentence-level paraphrase/translation
    narrative_like = text_len >= 18 and (punct_hits >= 1 or text_len >= 30)
    paraphrase_like = narrative_like and not has_hard_ann and gloss_hits == 0

    reflective_enough = (
        (core_kw_hits >= 1 and i_score >= 2 and text_len >= 20 and gloss_hits == 0)
        or (i_score >= 2 and text_len >= 22 and gloss_hits == 0 and marker_hits == 0)
        or paraphrase_like
        or (
            default == "interpretation"
            and i_score >= 1
            and text_len >= 16
            and gloss_hits == 0
            and marker_hits == 0
        )
    )
    if not reflective_enough:
        return "annotation"

    if default == "annotation":
        if i_score >= a_score + 3 and i_score >= 6 and core_kw_hits >= 2:
            return "interpretation"
        return "annotation"

    if a_score >= i_score + 2:
        return "annotation"
    return "interpretation"


def merge_interpretation_segments(cells: List[Tuple[int, str]]) -> List[Tuple[int, int, str]]:
    if not cells:
        return []

    cells = sorted(cells, key=lambda x: x[0])
    segments: List[Tuple[int, int, str]] = []
    start_sid, prev_sid = cells[0][0], cells[0][0]
    texts = [cells[0][1]]

    for sid, txt in cells[1:]:
        if sid == prev_sid + 1:
            texts.append(txt)
            prev_sid = sid
            continue
        segments.append((start_sid, prev_sid, "\n".join(texts)))
        start_sid = sid
        prev_sid = sid
        texts = [txt]

    segments.append((start_sid, prev_sid, "\n".join(texts)))
    return segments


def clean_zh(s: str) -> str:
    return re.sub(r"[^\u4e00-\u9fffA-Za-z0-9]", "", s or "")


def sentence_relevance(sentence: str, interpretation: str) -> int:
    """Lexical relevance between one source clause and interpretation text."""
    sent = clean_zh(sentence)
    text = interpretation or ""
    if len(sent) < 2:
        return 0

    score = 0
    if sent in text:
        score += 8

    # n-gram overlap: rewards explicit mention of source wording.
    for n, w in ((4, 3), (3, 2), (2, 1)):
        if len(sent) < n:
            continue
        grams = {sent[i : i + n] for i in range(len(sent) - n + 1)}
        hits = sum(1 for g in grams if g in text)
        score += min(4, hits) * w
    return score


def target_span_from_text(text: str) -> int:
    l = len(text or "")
    if l <= 40:
        return 3
    if l <= 90:
        return 4
    if l <= 160:
        return 6
    if l <= 260:
        return 8
    if l <= 420:
        return 10
    return 12


def smart_expand_interpretation_ranges(
    rows: List[List], sentence_text_by_text: Dict[int, Dict[int, str]], max_sid_by_text: Dict[int, int]
) -> List[List]:
    """Expand interpretation ranges with variable span using lexical cues + local boundaries.

    Keeps annotation untouched and only modifies interpretation start/end.
    """
    grouped: Dict[Tuple[int, str, str], List[List]] = {}
    for row in rows:
        key = (row[0], row[3], row[4])  # (text_id, commentator, dynasty)
        grouped.setdefault(key, []).append(row)

    out: List[List] = []

    for (text_id, commentator, dynasty), arr in grouped.items():
        arr = sorted(arr, key=lambda x: (x[1], x[2]))
        max_sid = max_sid_by_text[text_id]
        sent_map = sentence_text_by_text[text_id]

        centers = [((x[1] + x[2]) // 2) for x in arr]

        for idx, row in enumerate(arr):
            _, start_sid, end_sid, _, _, interp_text = row
            center = (start_sid + end_sid) // 2

            left_limit = 1
            right_limit = max_sid

            # Soft boundary: do not cross neighboring anchors, but allow overlap around them.
            if idx > 0:
                prev_center = centers[idx - 1]
                left_limit = max(left_limit, prev_center + 1)
            if idx + 1 < len(arr):
                next_center = centers[idx + 1]
                right_limit = min(right_limit, next_center - 1)

            # Preserve anchor sentence(s) for this interpretation.
            left_limit = min(left_limit, start_sid)
            right_limit = max(right_limit, end_sid)

            left = max(start_sid, left_limit)
            right = min(end_sid, right_limit)
            if left > right:
                left = right = start_sid

            rel = {}
            for sid in range(left_limit, right_limit + 1):
                rel[sid] = sentence_relevance(sent_map.get(sid, ""), interp_text)

            target_span = min(right_limit - left_limit + 1, max(2, target_span_from_text(interp_text)))
            max_span = min(right_limit - left_limit + 1, max(target_span + 5, target_span))

            # Phase 1: include adjacent clauses with positive lexical evidence.
            while True:
                l_score = rel.get(left - 1, -1) if left > left_limit else -1
                r_score = rel.get(right + 1, -1) if right < right_limit else -1
                best = max(l_score, r_score)
                if best <= 0:
                    break
                if (right - left + 1) >= max_span and best < 3:
                    break
                if r_score > l_score:
                    right += 1
                else:
                    left -= 1

            # Phase 2: if still short, smooth-expand to target span by local context.
            while (right - left + 1) < target_span:
                if left <= left_limit and right >= right_limit:
                    break
                l_ctx = (rel.get(left - 1, 0) + rel.get(left - 2, 0)) if left > left_limit else -1
                r_ctx = (rel.get(right + 1, 0) + rel.get(right + 2, 0)) if right < right_limit else -1
                if r_ctx > l_ctx and right < right_limit:
                    right += 1
                elif left > left_limit:
                    left -= 1
                elif right < right_limit:
                    right += 1
                else:
                    break

            out.append([text_id, left, right, commentator, dynasty, interp_text])

    return out


def detect_sentence_columns(ws, header_row: int) -> List[int]:
    """Detect sentence columns robustly even when the whole table is shifted."""
    meta_cols = []
    for col in range(1, ws.max_column + 1):
        v = normalize(ws.cell(header_row, col).value)
        if v in META_HEADER_LABELS:
            meta_cols.append(col)

    if meta_cols:
        start_col = max(meta_cols) + 1
    else:
        # Conservative fallback for legacy layout.
        start_col = 6

    sentence_cols = []
    for col in range(start_col, ws.max_column + 1):
        header = normalize(ws.cell(header_row, col).value)
        if header:
            sentence_cols.append(col)
    return sentence_cols


def resolve_layout_rows(ws, fallback_header_row: int, fallback_data_start: int) -> Tuple[int, int]:
    """Detect actual header/data rows.

    Supports variants where row 1 is only chapter title and row 2 has field headers.
    """
    meta_need = {"朝代", "书名"}
    for r in range(1, min(8, ws.max_row + 1)):
        row_vals = {normalize(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 40) + 1)}
        row_vals.discard("")
        # Need at least core meta headers and one role header.
        if meta_need.issubset(row_vals) and ("注者" in row_vals or "负责人" in row_vals or "分工" in row_vals):
            return r, r + 1
    return fallback_header_row, fallback_data_start


def detect_meta_columns(ws, header_row: int, data_start: int, first_sentence_col: int) -> Tuple[int, int, int]:
    """Return (commentator_col, dynasty_col, book_col)."""
    header_to_col: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = normalize(ws.cell(header_row, col).value)
        if v and v not in header_to_col:
            header_to_col[v] = col

    dynasty_col = header_to_col.get("朝代", 3)
    book_col = header_to_col.get("书名", 4)

    if "注者" in header_to_col:
        return header_to_col["注者"], dynasty_col, book_col

    # No explicit 注者 header (e.g. Sheet2): infer commentator column from data density/diversity.
    best_col = 1
    best_key = (-1, -1, -1)  # (unique_cnt, non_empty_cnt, -col)
    ignored_headers = {"分工", "负责人", "朝代", "书名", "特点"}

    for col in range(1, max(1, first_sentence_col)):
        if col in {dynasty_col, book_col}:
            continue
        h = normalize(ws.cell(header_row, col).value)
        if h in ignored_headers:
            continue

        values: List[str] = []
        for r in range(data_start, ws.max_row + 1):
            v = normalize(ws.cell(r, col).value)
            if v:
                values.append(v)

        if not values:
            continue

        key = (len(set(values)), len(values), -col)
        if key > best_key:
            best_key = key
            best_col = col

    return best_col, dynasty_col, book_col


def rebuild(raw_path: str, template_path: str, out_path: str, sync_to_raw: bool = True) -> None:
    raw_wb = openpyxl.load_workbook(raw_path, data_only=True)
    out_wb = openpyxl.load_workbook(template_path)

    ws_small = out_wb["small_sentences"]
    ws_ann = out_wb["annotations"]
    ws_int = out_wb["interpretations"]

    for ws in (ws_small, ws_ann, ws_int):
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    sentence_maps: Dict[str, List[Tuple[int, int, str]]] = {}
    max_sid_by_text: Dict[int, int] = {}
    sentence_text_by_text: Dict[int, Dict[int, str]] = {}
    layout_rows: Dict[str, Tuple[int, int]] = {}

    for cfg in TEXT_CONFIGS:
        ws = raw_wb[cfg.sheet]
        actual_header_row, actual_data_start = resolve_layout_rows(ws, cfg.header_row, cfg.data_start)
        layout_rows[cfg.sheet] = (actual_header_row, actual_data_start)
        sentence_maps[cfg.sheet] = []
        sid = 1
        sentence_cols = detect_sentence_columns(ws, actual_header_row)
        sentence_text_by_text[cfg.text_id] = {}
        for col in sentence_cols:
            sentence = normalize(ws.cell(actual_header_row, col).value)
            if not sentence:
                continue
            sentence_maps[cfg.sheet].append((col, sid, sentence))
            sentence_text_by_text[cfg.text_id][sid] = sentence
            ws_small.append([cfg.text_id, cfg.text_title, sid, sentence])
            sid += 1
        max_sid_by_text[cfg.text_id] = sid - 1

    annotation_rows: List[List] = []
    interpretation_rows: List[List] = []

    for cfg in TEXT_CONFIGS:
        ws = raw_wb[cfg.sheet]
        actual_header_row, actual_data_start = layout_rows[cfg.sheet]
        headers = sentence_maps[cfg.sheet]
        first_sentence_col = headers[0][0] if headers else ws.max_column + 1
        commentator_col, dynasty_col, book_col = detect_meta_columns(
            ws, actual_header_row, actual_data_start, first_sentence_col
        )

        for r in range(actual_data_start, ws.max_row + 1):
            commentator = normalize(ws.cell(r, commentator_col).value)
            if not commentator:
                continue
            dynasty = normalize(ws.cell(r, dynasty_col).value)
            book = normalize(ws.cell(r, book_col).value)

            row_interp_cells: List[Tuple[int, str]] = []
            for col, sid, _ in headers:
                text = normalize(ws.cell(r, col).value)
                if not text:
                    continue
                units = split_cell_by_sentence(text)
                interp_units: List[str] = []
                for unit in units:
                    kind = classify_cell(cfg.sheet, r, book, commentator, unit)
                    if kind == "annotation":
                        annotation_rows.append([cfg.text_id, sid, commentator, dynasty, unit])
                    else:
                        interp_units.append(unit)

                if interp_units:
                    merged_unit_text = "".join(interp_units)
                    row_interp_cells.append((sid, merged_unit_text))

            for start_sid, end_sid, merged_text in merge_interpretation_segments(row_interp_cells):
                interpretation_rows.append([cfg.text_id, start_sid, end_sid, commentator, dynasty, merged_text])

    interpretation_rows = smart_expand_interpretation_ranges(
        interpretation_rows, sentence_text_by_text, max_sid_by_text
    )

    for idx, row in enumerate(annotation_rows, start=1):
        ws_ann.append([idx, *row])
    for idx, row in enumerate(interpretation_rows, start=1):
        ws_int.append([idx, *row])

    out_wb.save(out_path)

    if not sync_to_raw:
        return

    dst_wb = openpyxl.load_workbook(raw_path)
    src_wb = openpyxl.load_workbook(out_path, data_only=True)

    for name in ["small_sentences", "annotations", "interpretations"]:
        if name in dst_wb.sheetnames:
            ws = dst_wb[name]
            if ws.max_row:
                ws.delete_rows(1, ws.max_row)
        else:
            ws = dst_wb.create_sheet(title=name)
        src_ws = src_wb[name]
        for row in src_ws.iter_rows(
            min_row=1,
            max_row=src_ws.max_row,
            min_col=1,
            max_col=src_ws.max_column,
            values_only=True,
        ):
            ws.append(list(row))

    # Keep generated sheets at the front.
    for name in ["interpretations", "annotations", "small_sentences"]:
        ws = dst_wb[name]
        dst_wb._sheets.remove(ws)
        dst_wb._sheets.insert(0, ws)

    dst_wb.save(raw_path)


if __name__ == "__main__":
    rebuild("zz.xlsx", "template.xlsx", "zz_structured.xlsx", sync_to_raw=True)
