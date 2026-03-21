from __future__ import annotations

import json
import re
from datetime import datetime, timezone

import openpyxl

SOURCE_FILE = "zz_structured.xlsx"
OUTPUT_FILE = "data.js"


def to_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    return int(float(text))


def normalize_cell_text(value: object) -> str:
    if value is None:
        return ""
    raw = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not raw:
        return ""
    return raw


def normalize_sentence_text(value: object) -> str:
    raw = normalize_cell_text(value)
    if not raw:
        return ""
    # Treat slash as paragraph delimiter from spreadsheet cells.
    text = re.sub(r"[\\/／]+", "\n", raw)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def sentence_break_after(value: object) -> bool:
    raw = normalize_cell_text(value)
    if not raw:
        return False
    return bool(re.search(r"[\\/／]\s*$", raw))


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[\s_]+", "", str(value).strip().lower())


def build_header_index(ws) -> dict[str, int]:
    index: dict[str, int] = {}
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for col, value in enumerate(header_row, start=1):
        key = normalize_header(value)
        if key and key not in index:
            index[key] = col
    return index


def pick_col(header_index: dict[str, int], candidates: list[str], default_col: int) -> int:
    for name in candidates:
        key = normalize_header(name)
        if key in header_index:
            return header_index[key]
    return default_col


def cell(row: tuple[object, ...], col: int) -> object:
    if col <= 0 or col > len(row):
        return None
    return row[col - 1]


def is_dynasty_like(text: str) -> bool:
    return bool(
        re.search(r"先秦|战国|戰國|秦|汉|漢|魏|晋|晉|隋|唐|宋|元|明|清|民国|民國|现代|現代|南朝|北朝", text)
    )


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    ws_sent = wb["small_sentences"]
    ws_anno = wb["annotations"]
    ws_interp = wb["interpretations"]

    sent_header = build_header_index(ws_sent)
    anno_header = build_header_index(ws_anno)
    interp_header = build_header_index(ws_interp)

    sent_text_id_col = pick_col(sent_header, ["text_id", "textid", "篇章id"], 1)
    sent_text_title_col = pick_col(sent_header, ["text_title", "texttitle", "篇章名", "标题", "標題"], 2)
    sent_sentence_id_col = pick_col(sent_header, ["sentence_id", "sentenceid", "句id", "分句id"], 3)
    sent_sentence_col = pick_col(sent_header, ["sentence", "原文", "分句", "句子"], 4)

    anno_id_col = pick_col(anno_header, ["annotation_id", "annotationid", "id", "编号", "編號", "序号", "序號"], 1)
    anno_text_id_col = pick_col(anno_header, ["text_id", "textid", "篇章id"], 2)
    anno_sentence_id_col = pick_col(anno_header, ["sentence_id", "sentenceid", "句id", "分句id"], 3)
    anno_commentator_col = pick_col(anno_header, ["commentator", "注者", "作者"], 4)
    anno_dynasty_col = pick_col(anno_header, ["dynasty", "朝代"], 5)
    anno_content_col = pick_col(anno_header, ["annotation", "注释", "註釋", "字义", "字義"], 6)

    interp_id_col = pick_col(
        interp_header, ["interpretation_id", "interpretationid", "id", "编号", "編號", "序号", "序號"], 1
    )
    interp_text_id_col = pick_col(interp_header, ["text_id", "textid", "篇章id"], 2)
    interp_start_col = pick_col(interp_header, ["start_sentence_id", "startsentenceid", "起始句id"], 3)
    interp_end_col = pick_col(interp_header, ["end_sentence_id", "endsentenceid", "结束句id", "結束句id"], 4)
    interp_commentator_col = pick_col(interp_header, ["commentator", "注者", "作者"], 5)
    interp_dynasty_col = pick_col(interp_header, ["dynasty", "朝代"], 6)
    interp_content_col = pick_col(interp_header, ["interpretation", "content", "阐释", "闡釋", "解释", "解釋"], 7)

    texts: dict[int, dict] = {}
    sentence_note_keys: set[tuple[int, int]] = set()

    annotations_by_key: dict[str, list[dict]] = {}
    interpretations: list[dict] = []
    interpretation_seen_keys: set[tuple[int, int, int, str, str, str]] = set()
    interpretation_dedup_count = 0

    for row in ws_anno.iter_rows(min_row=2, values_only=True):
        annotation_id = cell(row, anno_id_col)
        text_id = cell(row, anno_text_id_col)
        sentence_id = cell(row, anno_sentence_id_col)
        commentator_raw = cell(row, anno_commentator_col)
        dynasty_raw = cell(row, anno_dynasty_col)
        annotation_raw = cell(row, anno_content_col)

        tid = to_int(text_id)
        sid = to_int(sentence_id)
        commentator_text = str(commentator_raw or "未署名").strip()
        dynasty_text = str(dynasty_raw or "未详").strip()
        annotation_text = normalize_cell_text(annotation_raw)
        next_cell_text = normalize_cell_text(cell(row, anno_content_col + 1))

        # Compatibility fix for sheets where an extra column was inserted before commentator.
        # Pattern: commentator=数字序号, dynasty=注者, annotation=朝代, next_col=正文.
        if commentator_text.isdigit() and next_cell_text and is_dynasty_like(annotation_text):
            commentator_text = str(dynasty_raw or commentator_text).strip()
            dynasty_text = annotation_text or "未详"
            annotation_text = next_cell_text

        if tid is None or sid is None or not annotation_text:
            continue

        key = f"{tid}-{sid}"
        sentence_note_keys.add((tid, sid))
        annotations_by_key.setdefault(key, []).append(
            {
                "annotation_id": to_int(annotation_id),
                "commentator": commentator_text or "未署名",
                "dynasty": dynasty_text or "未详",
                "content": annotation_text,
            }
        )

    for row in ws_interp.iter_rows(min_row=2, values_only=True):
        interpretation_id = cell(row, interp_id_col)
        text_id = cell(row, interp_text_id_col)
        start_sid = cell(row, interp_start_col)
        end_sid = cell(row, interp_end_col)
        commentator = cell(row, interp_commentator_col)
        dynasty = cell(row, interp_dynasty_col)
        content = cell(row, interp_content_col)

        tid = to_int(text_id)
        start_id = to_int(start_sid)
        end_id = to_int(end_sid)
        content_text = normalize_cell_text(content)
        if tid is None or start_id is None or end_id is None or not content_text:
            continue

        if end_id < start_id:
            start_id, end_id = end_id, start_id

        item_commentator = str(commentator or "未署名")
        item_dynasty = str(dynasty or "未详")
        dedup_key = (tid, start_id, end_id, item_commentator, item_dynasty, content_text)
        if dedup_key in interpretation_seen_keys:
            interpretation_dedup_count += 1
            continue
        interpretation_seen_keys.add(dedup_key)

        for sid in range(start_id, end_id + 1):
            sentence_note_keys.add((tid, sid))

        interpretations.append(
            {
                "interpretation_id": to_int(interpretation_id),
                "text_id": tid,
                "start_sentence_id": start_id,
                "end_sentence_id": end_id,
                "commentator": item_commentator,
                "dynasty": item_dynasty,
                "content": content_text,
            }
        )

    for row in ws_sent.iter_rows(min_row=2, values_only=True):
        text_id = cell(row, sent_text_id_col)
        text_title = cell(row, sent_text_title_col)
        sentence_id = cell(row, sent_sentence_id_col)
        sentence = cell(row, sent_sentence_col)

        tid = to_int(text_id)
        sid = to_int(sentence_id)
        sentence_text = normalize_sentence_text(sentence)
        break_after = sentence_break_after(sentence)
        if tid is None or sid is None or not sentence_text:
            continue

        text_data = texts.setdefault(
            tid,
            {
                "text_id": tid,
                "text_title": str(text_title or f"篇章{tid}"),
                "sentences": [],
            },
        )

        text_data["sentences"].append(
            {
                "sentence_id": sid,
                "sentence": sentence_text,
                "break_after": break_after,
                "has_note": (tid, sid) in sentence_note_keys,
            }
        )

    text_list = []
    for _, text_data in sorted(texts.items(), key=lambda item: item[0]):
        text_data["sentences"].sort(key=lambda s: s["sentence_id"])
        text_list.append(text_data)

    payload = {
        "meta": {
            "source": SOURCE_FILE,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "text_count": len(text_list),
            "sentence_count": sum(len(item["sentences"]) for item in text_list),
            "annotation_key_count": len(annotations_by_key),
            "interpretation_count": len(interpretations),
            "interpretation_deduplicated": interpretation_dedup_count,
        },
        "texts": text_list,
        "annotations_by_key": annotations_by_key,
        "interpretations": interpretations,
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write("window.ZZ_DATA = ")
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")

    print(
        "Generated",
        OUTPUT_FILE,
        f"with {payload['meta']['sentence_count']} sentences and",
        f"{payload['meta']['interpretation_count']} interpretations.",
        f"Deduplicated {payload['meta']['interpretation_deduplicated']} interpretation rows.",
    )


if __name__ == "__main__":
    main()
