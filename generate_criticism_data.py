from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

CRITICISM_SOURCE = "criticism.xlsx"
TEXT_SOURCE = "zz_structured.xlsx"
DICT_SOURCE = "dic.xlsx"
OUTPUT_FILE = "criticism_data.js"
TEXT_DISPLAY_ORDER = [2, 3, 1]
TEXT_ORDER_INDEX = {tid: idx for idx, tid in enumerate(TEXT_DISPLAY_ORDER)}


def to_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    return int(float(text))


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[\s_]+", "", str(value).strip().lower())


def build_header_index(ws, max_scan: int = 200) -> dict[str, int]:
    index: dict[str, int] = {}
    limit = min(ws.max_column, max_scan)
    for col in range(1, limit + 1):
        key = normalize_header(ws.cell(1, col).value)
        if key and key not in index:
            index[key] = col
    return index


def pick_col(header_index: dict[str, int], candidates: list[str], default_col: int | None = None) -> int | None:
    for name in candidates:
        key = normalize_header(name)
        if key in header_index:
            return header_index[key]
    return default_col


def cell(ws, row: int, col: int | None) -> Any:
    if not col or col <= 0:
        return None
    return ws.cell(row, col).value


def group_token(raw: object) -> str:
    text = normalize_text(raw)
    if not text:
        return ""
    m = re.search(r"(\d+)", text)
    if m:
        return m.group(1)
    return text


def split_highlight_terms(raw: object) -> list[str]:
    text = normalize_text(raw)
    if not text:
        return []
    terms: list[str] = []
    seen: set[str] = set()
    for part in re.split(r"[\n\r,，、;；/|]+", text):
        term = part.strip()
        if not term or term in seen:
            continue
        seen.add(term)
        terms.append(term)
    return terms


def load_variant_mappings() -> list[dict[str, object]]:
    if not Path(DICT_SOURCE).exists():
        return []

    wb = openpyxl.load_workbook(DICT_SOURCE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = build_header_index(ws, max_scan=80)

    canonical_col = pick_col(
        headers,
        ["选取的字", "选取字", "标准字", "展示字", "呈现字", "canonical", "display", "target"],
        1,
    )
    variant_col = pick_col(
        headers,
        ["可能的字", "可能字", "异文", "異文", "他本", "其它版本", "其他版本", "variant", "alternate"],
        2,
    )
    text_id_col = pick_col(headers, ["text_id", "textid", "篇章id", "书"], None)
    sentence_id_col = pick_col(headers, ["sentence_id", "sentenceid", "句id", "分句id"], None)

    fallback_scope_kind = ""
    if not text_id_col and not sentence_id_col and ws.max_column >= 3:
        third_header_raw = normalize_text(ws.cell(1, 3).value)
        third_header_norm = normalize_header(third_header_raw)
        if "sentence" in third_header_norm or "句" in third_header_raw:
            fallback_scope_kind = "sentence"
        elif "text" in third_header_norm or "篇" in third_header_raw or "书" in third_header_raw:
            fallback_scope_kind = "text"

    out: list[dict[str, object]] = []
    seen: set[tuple[str, str, int | None, int | None]] = set()
    for row in range(2, ws.max_row + 1):
        canonical = normalize_text(cell(ws, row, canonical_col))
        variant = normalize_text(cell(ws, row, variant_col))
        if not canonical or not variant or canonical == variant:
            continue

        text_id = to_int(cell(ws, row, text_id_col))
        sentence_id = to_int(cell(ws, row, sentence_id_col))
        if text_id is None and sentence_id is None and fallback_scope_kind:
            scope_value = to_int(cell(ws, row, 3))
            if fallback_scope_kind == "text":
                text_id = scope_value
            elif fallback_scope_kind == "sentence":
                sentence_id = scope_value

        key = (canonical, variant, text_id, sentence_id)
        if key in seen:
            continue
        seen.add(key)

        item: dict[str, object] = {
            "canonical": canonical,
            "variant": variant,
        }
        if text_id is not None:
            item["text_id"] = text_id
        if sentence_id is not None:
            item["sentence_id"] = sentence_id
        out.append(item)

    return out


def load_sentence_map() -> tuple[dict[tuple[int, int], dict[str, object]], dict[int, str]]:
    wb = openpyxl.load_workbook(TEXT_SOURCE, data_only=True)
    ws = wb["small_sentences"]
    headers = build_header_index(ws, max_scan=80)

    text_id_col = pick_col(headers, ["text_id", "textid", "篇章id"], 1)
    title_col = pick_col(headers, ["text_title", "texttitle", "篇章名", "标题", "標題"], 2)
    sentence_id_col = pick_col(headers, ["sentence_id", "sentenceid", "句id", "分句id"], 3)
    sentence_col = pick_col(headers, ["sentence", "原文", "分句", "句子"], 4)

    sentence_map: dict[tuple[int, int], dict[str, object]] = {}
    title_map: dict[int, str] = {}

    for row in range(2, ws.max_row + 1):
        text_id = to_int(cell(ws, row, text_id_col))
        sentence_id = to_int(cell(ws, row, sentence_id_col))
        if text_id is None or sentence_id is None:
            continue

        text_title = normalize_text(cell(ws, row, title_col)) or f"篇章{text_id}"
        sentence = normalize_text(cell(ws, row, sentence_col))

        title_map[text_id] = text_title
        sentence_map[(text_id, sentence_id)] = {
            "text_id": text_id,
            "text_title": text_title,
            "sentence_id": sentence_id,
            "sentence": sentence,
        }

    return sentence_map, title_map


def load_relations_and_refs() -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    wb = openpyxl.load_workbook(CRITICISM_SOURCE, data_only=True)

    # Preferred normalized layout.
    if "eval_edges" in wb.sheetnames and "ref_notes" in wb.sheetnames:
        ws_e = wb["eval_edges"]
        ws_r = wb["ref_notes"]

        h_e = build_header_index(ws_e)
        e_event = pick_col(h_e, ["event_id", "eventid"], 2)
        e_text = pick_col(h_e, ["text_id", "textid", "书"], 3)
        e_sent = pick_col(h_e, ["sentence_id", "sentenceid", "id"], 4)
        e_from_author = pick_col(h_e, ["from_author", "fromauthor", "人"], 5)
        e_from_text = pick_col(h_e, ["from_text", "fromtext", "话"], 6)
        e_to_author = pick_col(h_e, ["to_author", "toauthor", "评价者"], 7)
        e_eval_text = pick_col(h_e, ["eval_text", "evaltext", "评价语"], 8)
        e_stance = pick_col(h_e, ["stance", "态度"], 9)
        e_highlight = pick_col(h_e, ["highlight", "highlights", "原文高亮", "高亮"], 10)

        relations: list[dict[str, object]] = []
        for row in range(2, ws_e.max_row + 1):
            text_id = to_int(cell(ws_e, row, e_text))
            sentence_id = to_int(cell(ws_e, row, e_sent))
            from_author = normalize_text(cell(ws_e, row, e_from_author))
            from_text = normalize_text(cell(ws_e, row, e_from_text))
            to_author = normalize_text(cell(ws_e, row, e_to_author))
            eval_text = normalize_text(cell(ws_e, row, e_eval_text))
            stance = normalize_text(cell(ws_e, row, e_stance)).upper()
            highlight = normalize_text(cell(ws_e, row, e_highlight))
            group_id = normalize_text(cell(ws_e, row, e_event))

            if text_id is None or sentence_id is None:
                continue
            if not from_author and not from_text and not to_author and not eval_text:
                continue
            if stance not in {"Y", "N"}:
                continue
            if not to_author or not eval_text:
                continue

            relations.append(
                {
                    "group_id": group_id,
                    "group_token": group_token(group_id),
                    "text_id": text_id,
                    "sentence_id": sentence_id,
                    "from_author": from_author,
                    "from_text": from_text,
                    "to_author": to_author,
                    "eval_text": eval_text,
                    "stance": stance,
                    "highlight": highlight,
                }
            )

        h_r = build_header_index(ws_r)
        r_group = pick_col(h_r, ["ref_group_id", "refgroupid", "group_id", "groupid"], 2)
        r_text = pick_col(h_r, ["text_id", "textid", "书"], 3)
        r_sent = pick_col(h_r, ["sentence_id", "sentenceid", "id"], 4)
        r_author = pick_col(h_r, ["ref_author", "refauthor", "from_author", "人"], 5)
        r_text_col = pick_col(h_r, ["ref_text", "reftext", "from_text", "话"], 6)

        refs: list[dict[str, object]] = []
        for row in range(2, ws_r.max_row + 1):
            text_id = to_int(cell(ws_r, row, r_text))
            sentence_id = to_int(cell(ws_r, row, r_sent))
            ref_author = normalize_text(cell(ws_r, row, r_author))
            ref_text = normalize_text(cell(ws_r, row, r_text_col))
            group_id = normalize_text(cell(ws_r, row, r_group))

            if text_id is None or sentence_id is None:
                continue
            if not ref_author and not ref_text:
                continue

            refs.append(
                {
                    "group_id": group_id,
                    "group_token": group_token(group_id),
                    "text_id": text_id,
                    "sentence_id": sentence_id,
                    "ref_author": ref_author,
                    "ref_text": ref_text,
                }
            )

        return relations, refs

    # Legacy single-sheet fallback.
    ws = None
    for candidate in wb.worksheets:
        if candidate.max_row > 1:
            ws = candidate
            break
    if ws is None:
        return [], []

    headers = build_header_index(ws)
    c_count = pick_col(headers, ["计数", "count"], 1)
    c_text = pick_col(headers, ["书", "text_id", "textid"], 2)
    c_sent = pick_col(headers, ["id", "sentence_id", "sentenceid"], 3)
    c_from_author = pick_col(headers, ["人", "from_author", "fromauthor"], 4)
    c_from_text = pick_col(headers, ["话", "from_text", "fromtext"], 5)
    c_to_author = pick_col(headers, ["评价者", "to_author", "toauthor"], 6)
    c_eval_text = pick_col(headers, ["评价语", "eval_text", "evaltext"], 7)
    c_stance = pick_col(headers, ["态度", "stance"], 8)

    relations: list[dict[str, object]] = []
    refs: list[dict[str, object]] = []

    for row in range(2, ws.max_row + 1):
        count = normalize_text(cell(ws, row, c_count))
        text_id = to_int(cell(ws, row, c_text))
        sentence_id = to_int(cell(ws, row, c_sent))
        from_author = normalize_text(cell(ws, row, c_from_author))
        from_text = normalize_text(cell(ws, row, c_from_text))
        to_author = normalize_text(cell(ws, row, c_to_author))
        eval_text = normalize_text(cell(ws, row, c_eval_text))
        stance = normalize_text(cell(ws, row, c_stance)).upper()

        if text_id is None or sentence_id is None:
            continue
        if not from_author and not from_text:
            continue

        if stance in {"Y", "N"} and to_author and eval_text:
            group_id = f"E{count}" if count else ""
            relations.append(
                {
                    "group_id": group_id,
                    "group_token": group_token(group_id or count),
                    "text_id": text_id,
                    "sentence_id": sentence_id,
                    "from_author": from_author,
                    "from_text": from_text,
                    "to_author": to_author,
                    "eval_text": eval_text,
                    "stance": stance,
                    "highlight": "",
                }
            )
        else:
            ref_group_id = f"G{count}" if count else ""
            refs.append(
                {
                    "group_id": ref_group_id,
                    "group_token": group_token(ref_group_id or count),
                    "text_id": text_id,
                    "sentence_id": sentence_id,
                    "ref_author": from_author,
                    "ref_text": from_text,
                }
            )

    return relations, refs


def main() -> None:
    sentence_map, title_map = load_sentence_map()
    relations, refs = load_relations_and_refs()
    variant_mappings = load_variant_mappings()

    groups: dict[tuple[int, int, str], dict[str, object]] = {}

    def ensure_group(text_id: int, sentence_id: int, token: str) -> dict[str, object]:
        key = (text_id, sentence_id, token)
        if key in groups:
            return groups[key]

        sent = sentence_map.get((text_id, sentence_id), {})
        group_id = f"G{token}" if token else ""
        group = {
            "group_token": token,
            "group_id": group_id,
            "text_id": text_id,
            "text_title": sent.get("text_title") or title_map.get(text_id) or f"篇章{text_id}",
            "sentence_id": sentence_id,
            "sentence": sent.get("sentence") or "",
            "highlights": [],
            "relations": [],
            "references": [],
        }
        groups[key] = group
        return group

    for row in relations:
        token = normalize_text(row.get("group_token"))
        group = ensure_group(int(row["text_id"]), int(row["sentence_id"]), token)
        relation_group_id = normalize_text(row.get("group_id"))
        if relation_group_id:
            group["group_id"] = relation_group_id
        for term in split_highlight_terms(row.get("highlight")):
            if term not in group["highlights"]:
                group["highlights"].append(term)
        group["relations"].append(
            {
                "from_author": normalize_text(row.get("from_author")),
                "from_text": normalize_text(row.get("from_text")),
                "to_author": normalize_text(row.get("to_author")),
                "eval_text": normalize_text(row.get("eval_text")),
                "stance": normalize_text(row.get("stance")).upper(),
                "highlight": normalize_text(row.get("highlight")),
            }
        )

    for row in refs:
        token = normalize_text(row.get("group_token"))
        group = ensure_group(int(row["text_id"]), int(row["sentence_id"]), token)
        ref_group_id = normalize_text(row.get("group_id"))
        # Keep event-group label when relations already exist.
        if ref_group_id and not group["relations"] and not normalize_text(group.get("group_id")):
            group["group_id"] = ref_group_id
        group["references"].append(
            {
                "ref_author": normalize_text(row.get("ref_author")),
                "ref_text": normalize_text(row.get("ref_text")),
            }
        )

    # Sort and pack by text.
    by_text: dict[int, dict[str, object]] = {}

    def token_sort_key(token: str) -> tuple[int, int | str]:
        m = re.fullmatch(r"\d+", token or "")
        if m:
            return (0, int(token))
        return (1, token or "")

    for group in groups.values():
        text_id = int(group["text_id"])
        bucket = by_text.setdefault(
            text_id,
            {
                "text_id": text_id,
                "text_title": group["text_title"],
                "groups": [],
            },
        )

        group["relations"].sort(
            key=lambda x: (
                x["from_author"],
                x["from_text"],
                x["to_author"],
                x["eval_text"],
                x["stance"],
            )
        )
        group["references"].sort(key=lambda x: (x["ref_author"], x["ref_text"]))
        bucket["groups"].append(group)

    text_list: list[dict[str, object]] = []
    for _, text in sorted(
        by_text.items(), key=lambda kv: (TEXT_ORDER_INDEX.get(kv[0], 10_000), kv[0])
    ):
        text["groups"].sort(
            key=lambda g: (
                int(g["sentence_id"]),
                token_sort_key(str(g.get("group_token") or "")),
            )
        )
        text["group_count"] = len(text["groups"])
        text["relation_count"] = sum(len(g["relations"]) for g in text["groups"])
        text["reference_count"] = sum(len(g["references"]) for g in text["groups"])
        text_list.append(text)

    payload = {
        "meta": {
            "source": CRITICISM_SOURCE,
            "text_source": TEXT_SOURCE,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "text_count": len(text_list),
            "group_count": sum(len(t["groups"]) for t in text_list),
            "relation_count": sum(t["relation_count"] for t in text_list),
            "reference_count": sum(t["reference_count"] for t in text_list),
            "variant_mapping_count": len(variant_mappings),
        },
        "texts": text_list,
        "variant_mappings": variant_mappings,
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write("window.CRITICISM_DATA = ")
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")

    print(
        "Generated",
        OUTPUT_FILE,
        f"with {payload['meta']['group_count']} groups,",
        f"{payload['meta']['relation_count']} relations and",
        f"{payload['meta']['reference_count']} references.",
        f"Loaded {payload['meta']['variant_mapping_count']} variant mappings.",
    )


if __name__ == "__main__":
    main()
