from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from typing import Any

import openpyxl

PHILOSOPHY_SOURCE = "philosophy.xlsx"
TEXT_SOURCE = "zz_structured.xlsx"
OUTPUT_FILE = "philosophy_data.js"
TEXT_DISPLAY_ORDER = [2, 3, 1]
TEXT_ORDER_INDEX = {tid: idx for idx, tid in enumerate(TEXT_DISPLAY_ORDER)}


def to_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    return int(float(text))


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[\s_]+", "", str(value).strip().lower())


def build_header_index(ws, max_scan: int = 200) -> dict[str, int]:
    index: dict[str, int] = {}
    limit = min(ws.max_column, max_scan)
    for col in range(1, limit + 1):
        key = normalize_header(ws.cell(1, col).value)
        if key and key not in index:
            index[key] = col
    return index


def pick_col(header_index: dict[str, int], candidates: list[str], default_col: int | None = None) -> int | None:
    for name in candidates:
        key = normalize_header(name)
        if key in header_index:
            return header_index[key]
    return default_col


def cell(ws, row: int, col: int | None) -> Any:
    if not col or col <= 0:
        return None
    return ws.cell(row, col).value


def find_note_columns(ws, max_scan: int = 120) -> list[int]:
    cols: list[int] = []
    limit = min(ws.max_column, max_scan)
    for col in range(1, limit + 1):
        raw = normalize_text(ws.cell(1, col).value)
        if not raw:
            continue
        normalized = normalize_header(raw)
        if raw.startswith("注") or normalized.startswith("note") or normalized.startswith("annotation"):
            cols.append(col)
    return cols


def load_text_title_map() -> dict[int, str]:
    wb = openpyxl.load_workbook(TEXT_SOURCE, data_only=True)
    if "small_sentences" not in wb.sheetnames:
        return {}

    ws = wb["small_sentences"]
    headers = build_header_index(ws, max_scan=80)
    text_id_col = pick_col(headers, ["text_id", "textid", "篇章id"], 1)
    title_col = pick_col(headers, ["text_title", "texttitle", "篇章名", "标题", "標題"], 2)

    title_map: dict[int, str] = {}
    for row in range(2, ws.max_row + 1):
        text_id = to_int(cell(ws, row, text_id_col))
        if text_id is None:
            continue
        text_title = normalize_text(cell(ws, row, title_col)) or f"篇章{text_id}"
        if text_id not in title_map:
            title_map[text_id] = text_title
    return title_map


def load_concepts() -> tuple[list[dict[str, object]], int]:
    wb = openpyxl.load_workbook(PHILOSOPHY_SOURCE, data_only=True)
    concepts: list[dict[str, object]] = []

    for sheet_index, ws in enumerate(wb.worksheets):
        headers = build_header_index(ws, max_scan=200)
        text_col = pick_col(headers, ["text_id", "textid", "篇章id", "书"], 1)
        concept_col = pick_col(headers, ["概念", "concept", "term", "词条"], 2)
        scope_col = pick_col(headers, ["内涵外延", "内涵", "外延", "释义", "定義", "定义", "说明", "explanation"], 3)
        note_cols = find_note_columns(ws)

        for row in range(2, ws.max_row + 1):
            text_id = to_int(cell(ws, row, text_col))
            concept = normalize_text(cell(ws, row, concept_col))
            scope = normalize_text(cell(ws, row, scope_col))
            notes = [normalize_text(cell(ws, row, col)) for col in note_cols]
            notes = [item for item in notes if item]

            if text_id is None:
                if not concept and not scope and not notes:
                    continue
                # Skip malformed non-empty rows without text_id.
                continue

            if not concept and not scope and not notes:
                continue

            concepts.append(
                {
                    "text_id": text_id,
                    "concept": concept,
                    "scope": scope,
                    "notes": notes,
                    "source_sheet": ws.title,
                    "source_sheet_index": sheet_index,
                    "source_row": row,
                }
            )

    return concepts, len(wb.worksheets)


def main() -> None:
    title_map = load_text_title_map()
    concepts, sheet_count = load_concepts()

    by_text: dict[int, dict[str, object]] = {}
    for idx, item in enumerate(concepts, start=1):
        text_id = int(item["text_id"])
        bucket = by_text.setdefault(
            text_id,
            {
                "text_id": text_id,
                "text_title": title_map.get(text_id) or f"篇章{text_id}",
                "concepts": [],
            },
        )

        concept_name = normalize_text(item.get("concept")) or f"未命名概念{idx}"
        bucket["concepts"].append(
            {
                "concept_id": f"{item['source_sheet']}-{item['source_row']}",
                "concept": concept_name,
                "scope": normalize_text(item.get("scope")),
                "notes": list(item.get("notes") or []),
                "source_sheet": normalize_text(item.get("source_sheet")),
                "source_sheet_index": int(item.get("source_sheet_index") or 0),
                "source_row": int(item.get("source_row") or 0),
            }
        )

    text_list: list[dict[str, object]] = []
    total_note_count = 0
    total_concept_count = 0

    for _, text in sorted(
        by_text.items(), key=lambda kv: (TEXT_ORDER_INDEX.get(kv[0], 10_000), kv[0])
    ):
        concepts_list = list(text["concepts"])
        concepts_list.sort(
            key=lambda x: (
                int(x.get("source_sheet_index") or 0),
                int(x.get("source_row") or 0),
            )
        )
        for concept in concepts_list:
            concept.pop("source_sheet_index", None)
            concept.pop("source_row", None)
        text["concepts"] = concepts_list
        text["concept_count"] = len(concepts_list)
        text["note_count"] = sum(len(c.get("notes", [])) for c in concepts_list)
        total_concept_count += int(text["concept_count"])
        total_note_count += int(text["note_count"])
        text_list.append(text)

    payload = {
        "meta": {
            "source": PHILOSOPHY_SOURCE,
            "text_source": TEXT_SOURCE,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "sheet_count": sheet_count,
            "text_count": len(text_list),
            "concept_count": total_concept_count,
            "note_count": total_note_count,
        },
        "texts": text_list,
    }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        f.write("window.PHILOSOPHY_DATA = ")
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")

    print(
        "Generated",
        OUTPUT_FILE,
        f"with {payload['meta']['text_count']} texts,",
        f"{payload['meta']['concept_count']} concepts and",
        f"{payload['meta']['note_count']} notes.",
    )


if __name__ == "__main__":
    main()
